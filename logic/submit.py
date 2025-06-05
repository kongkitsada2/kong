import re
from datetime import datetime
from flask import jsonify
import openpyxl
from utils.gsheet import get_entry_mapping_by_form_url
from urllib.parse import urlencode

EXCEL_PATH = "‡∏õ‡∏±‡πâ‡∏°‡∏•‡∏°.xlsx"

def normalize_model(text):
    return re.sub(r'[\s\-\(\)]', '', text.upper())

def check_and_deduct_new_form_data():
    models = get_all_models_from_last_response_file()
    update_last_response_file_if_needed(models)
    return {"status": "success", "message": "‡πÄ‡∏ä‡πá‡∏Ñ‡πÅ‡∏•‡∏∞‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏•‡πà‡∏≤‡∏™‡∏∏‡∏î‡πÅ‡∏•‡πâ‡∏ß"}

def submit_qr_logic(request):
    data = request.json
    qr_text = data.get("data", "").strip()
    update_af = data.get("update_af", True)

    if not qr_text:
        return jsonify({"status": "error", "message": "QR code ‡∏ß‡πà‡∏≤‡∏á‡πÄ‡∏õ‡∏•‡πà‡∏≤"}), 400

    lines = qr_text.split("\n")
    model_line = next((line for line in lines if "model" in line.lower()), None)
    if not model_line or ":" not in model_line:
        return jsonify({"status": "error", "message": "‡πÑ‡∏°‡πà‡∏û‡∏ö MODEL ‡πÉ‡∏ô QR"}), 400

    model_raw = model_line.split(":", 1)[-1].strip()
    model_text = normalize_model(model_raw)

    try:
        def safe_get(row, index):
            return row[index] if len(row) > index and row[index] is not None else ""

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        matched_row = next((row for row in ws.iter_rows(min_row=2, values_only=True)
                            if row and len(row) > 1 and normalize_model(str(row[1])) == model_text), None)

        if not matched_row:
            return jsonify({"status": "error", "message": f"‡πÑ‡∏°‡πà‡∏û‡∏ö Model '{model_raw}' ‡πÉ‡∏ô Excel"}), 404

        form_url = next((safe_get(matched_row, i) for i, header in enumerate(headers) if "form" in header.lower()), "")

        if not form_url:
            return jsonify({"status": "error", "message": "‡πÑ‡∏°‡πà‡∏û‡∏ö Form URL ‡πÉ‡∏ô Excel"}), 400

        ENTRY_MAPPING = get_entry_mapping_by_form_url(form_url)
        form_data = {}

        if "‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà" in ENTRY_MAPPING:
            form_data[ENTRY_MAPPING["‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà"]] = datetime.now().strftime("%Y-%m-%d")

        for col_index, header in enumerate(headers):
            entry_id = ENTRY_MAPPING.get(header)
            if entry_id:
                value = safe_get(matched_row, col_index)
                form_data[entry_id] = value.strip().upper() if header in ["MODEL", "‡∏¢‡∏µ‡πà‡∏´‡πâ‡∏≠‡∏õ‡∏±‡πä‡∏°‡∏•‡∏°"] else value

        query_string = urlencode(form_data)

        # üîç ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡πà‡∏ß‡∏ô‡∏ô‡∏µ‡πâ
        form_date = form_data.get(ENTRY_MAPPING["‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà"])
        today_str = datetime.now().strftime("%Y-%m-%d")

        if form_date != today_str:
            update_last_response_file_only(model_text)
            return jsonify({"status": "info", "message": "‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πà‡∏Ç‡∏≠‡∏á‡∏ß‡∏±‡∏ô‡∏ô‡∏µ‡πâ, ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡πÑ‡∏ü‡∏•‡πå .txt ‡πÅ‡∏•‡πâ‡∏ß"})

        # üîÅ ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πà‡∏ß‡∏±‡∏ô‡∏ô‡∏µ‡πâ ‚Üí redirect ‡πÑ‡∏õ‡∏Å‡∏£‡∏≠‡∏Å‡∏ü‡∏≠‡∏£‡πå‡∏°
        redirect_url = f"{form_url}?{query_string}"
        return jsonify({"status": "redirect", "url": redirect_url})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
